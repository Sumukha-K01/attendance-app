"""
Attendance Backup Utilities

This module provides utilities for backing up attendance data,
creating reports, and managing automated backups.
"""

import os
import json
import zipfile
from datetime import datetime, timedelta
from django.core.mail import EmailMessage
from django.conf import settings
from django.db.models import Q, Count
from attendance.models import Attendance, Student, Classroom, Houses
from accounts.models import Branch
import tempfile
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


class AttendanceBackupManager:
    """Manager class for handling attendance backups"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def create_full_backup(self, 
                          email: str = 'vinodrc3@gmail.com',
                          days_back: int = 30,
                          branch_name: Optional[str] = None) -> bool:
        """
        Create and send a full backup of attendance data as Excel file
        
        Args:
            email: Email address to send backup to
            days_back: Number of days back to include
            branch_name: Specific branch to backup (optional)
            
        Returns:
            bool: True if backup was successful, False otherwise
        """
        try:
            end_date = datetime.now().date()
            start_date = end_date - timedelta(days=days_back)
            
            excel_file = self._create_excel_backup(start_date, end_date, branch_name)
            
            if excel_file:
                self._send_backup_email(email, [excel_file], start_date, end_date)
                self._cleanup_files([excel_file])
                return True
            
            return False
            
        except Exception as e:
            print(f"Backup failed: {str(e)}")
            return False
    
    def create_summary_report(self, days_back: int = 7) -> Dict:
        """
        Create a summary report of attendance statistics
        
        Args:
            days_back: Number of days to include in summary
            
        Returns:
            Dict: Summary statistics
        """
        end_date = datetime.now().date()
        start_date = end_date - timedelta(days=days_back)
        
        # Get attendance data for the period
        attendance_data = Attendance.objects.filter(
            date__range=[start_date, end_date]
        ).select_related('student', 'student__classroom', 'student__branch')
        
        # Calculate statistics
        total_records = attendance_data.count()
        total_students = attendance_data.values('student').distinct().count()
        
        # Attendance type statistics
        attendance_types = ['morning_attendance', 'evening_class_attendance', 
                          'morning_pt_attendance', 'games_attendance', 'night_dorm_attendance']
        
        stats_by_type = {}
        for att_type in attendance_types:
            type_stats = attendance_data.values(att_type).annotate(count=Count(att_type))
            stats_by_type[att_type] = {item[att_type]: item['count'] for item in type_stats}
        
        # Branch-wise statistics
        branch_stats = attendance_data.values('student__branch__name').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return {
            'summary': {
                'date_range': f"{start_date} to {end_date}",
                'total_records': total_records,
                'total_students': total_students,
                'days_covered': days_back
            },
            'attendance_by_type': stats_by_type,
            'branch_statistics': list(branch_stats)
        }
    
    def schedule_daily_backup(self, email: str = 'vinodrc3@gmail.com') -> bool:
        """
        Schedule a daily backup (previous day's data)
        
        Args:
            email: Email to send backup to
            
        Returns:
            bool: Success status
        """
        return self.create_full_backup(
            email=email,
            days_back=1
        )
    
    def schedule_weekly_backup(self, email: str = 'vinodrc3@gmail.com') -> bool:
        """
        Schedule a weekly backup
        
        Args:
            email: Email to send backup to
            
        Returns:
            bool: Success status
        """
        return self.create_full_backup(
            email=email,
            days_back=7
        )
    
    def schedule_monthly_backup(self, email: str = 'vinodrc3@gmail.com') -> bool:
        """
        Schedule a monthly backup
        
        Args:
            email: Email to send backup to
            
        Returns:
            bool: Success status
        """
        return self.create_full_backup(
            email=email,
            days_back=30
        )
    
    def _create_excel_backup(self, start_date, end_date, branch_name=None) -> Optional[str]:
        """Create Excel backup file with attendance data"""
        
        # Get attendance data
        attendance_query = Attendance.objects.filter(
            date__range=[start_date, end_date]
        ).select_related('student', 'student__classroom', 'student__house', 'student__branch')
        
        if branch_name:
            try:
                branch = Branch.objects.get(name__iexact=branch_name)
                attendance_query = attendance_query.filter(student__branch=branch)
            except Branch.DoesNotExist:
                pass  # Include all branches if specified branch doesn't exist
        
        attendance_data = list(attendance_query.order_by('date', 'student__classroom', 'student__roll_number'))
        
        if not attendance_data:
            return None
        
        # Create Excel file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"JNV_Attendance_Backup_{start_date}_to_{end_date}_{timestamp}.xlsx"
        file_path = os.path.join(self.temp_dir, filename)
        
        try:
            workbook = openpyxl.Workbook()
            
            # Create main attendance sheet
            ws_main = workbook.active
            ws_main.title = "Attendance Data"
            
            # Headers matching the exact database table structure for easy restoration
            headers = [
                'id',                           # attendance.id (primary key)
                'student_id',                   # attendance.student_id (foreign key)
                'date',                         # attendance.date
                'morning_attendance',           # attendance.morning_attendance
                'evening_class_attendance',     # attendance.evening_class_attendance
                'morning_pt_attendance',        # attendance.morning_pt_attendance
                'games_attendance',             # attendance.games_attendance
                'night_dorm_attendance',        # attendance.night_dorm_attendance
                # Additional human-readable columns for reference
                'student_name',                 # For human readability
                'roll_number',                  # For human readability
                'class_name',                   # For human readability
                'house_name',                   # For human readability
                'branch_name'                   # For human readability
            ]
            
            # Apply header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Add data rows with exact database structure
            for row, record in enumerate(attendance_data, 2):
                # Database columns (for restoration)
                ws_main.cell(row=row, column=1, value=record.id)                                      # id
                ws_main.cell(row=row, column=2, value=record.student.id)                             # student_id
                ws_main.cell(row=row, column=3, value=record.date.strftime('%Y-%m-%d'))              # date
                ws_main.cell(row=row, column=4, value=record.morning_attendance or 'NOT_MARKED')     # morning_attendance
                ws_main.cell(row=row, column=5, value=record.evening_class_attendance or 'NOT_MARKED') # evening_class_attendance
                ws_main.cell(row=row, column=6, value=record.morning_pt_attendance or 'NOT_MARKED')  # morning_pt_attendance
                ws_main.cell(row=row, column=7, value=record.games_attendance or 'NOT_MARKED')       # games_attendance
                ws_main.cell(row=row, column=8, value=record.night_dorm_attendance or 'NOT_MARKED')  # night_dorm_attendance
                
                # Human-readable columns (for reference)
                ws_main.cell(row=row, column=9, value=record.student.name)                           # student_name
                ws_main.cell(row=row, column=10, value=record.student.roll_number)                   # roll_number
                ws_main.cell(row=row, column=11, value=record.student.classroom.name if record.student.classroom else 'N/A') # class_name
                ws_main.cell(row=row, column=12, value=record.student.house.get_name_display() if record.student.house else 'N/A') # house_name
                ws_main.cell(row=row, column=13, value=record.student.branch.name if record.student.branch else 'N/A') # branch_name
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws_main.column_dimensions[column_letter].width = 15
            
            # Create summary sheet
            ws_summary = workbook.create_sheet("Summary")
            
            # Summary data
            total_records = len(attendance_data)
            total_students = len(set(record.student.id for record in attendance_data))
            date_range = f"{start_date} to {end_date}"
            
            # Branch statistics
            branch_stats = {}
            for record in attendance_data:
                branch_name = record.student.branch.name if record.student.branch else 'Unknown'
                branch_stats[branch_name] = branch_stats.get(branch_name, 0) + 1
            
            # Class statistics
            class_stats = {}
            for record in attendance_data:
                class_name = record.student.classroom.name if record.student.classroom else 'Unknown'
                class_stats[class_name] = class_stats.get(class_name, 0) + 1
            
            # Add summary content
            summary_data = [
                ['JNV Attendance Backup Summary', ''],
                ['Generated On', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ['Date Range', date_range],
                ['Total Records', total_records],
                ['Total Students', total_students],
                ['', ''],
                ['Branch Statistics', ''],
                ['Branch Name', 'Record Count']
            ]
            
            for branch, count in sorted(branch_stats.items()):
                summary_data.append([branch, count])
            
            summary_data.extend([
                ['', ''],
                ['Class Statistics', ''],
                ['Class Name', 'Record Count']
            ])
            
            for class_name, count in sorted(class_stats.items()):
                summary_data.append([class_name, count])
            
            # Write summary data
            for row, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row, column=1, value=label)
                ws_summary.cell(row=row, column=2, value=value)
                
                # Style headers
                if label in ['JNV Attendance Backup Summary', 'Branch Statistics', 'Class Statistics']:
                    ws_summary.cell(row=row, column=1).font = Font(bold=True, size=14)
                elif label in ['Branch Name', 'Class Name']:
                    ws_summary.cell(row=row, column=1).font = Font(bold=True)
                    ws_summary.cell(row=row, column=2).font = Font(bold=True)
            
            # Auto-adjust summary column widths
            ws_summary.column_dimensions['A'].width = 25
            ws_summary.column_dimensions['B'].width = 15
            
            # Create Restoration Instructions Sheet
            ws_restore = workbook.create_sheet(title="Restoration Instructions")
            
            restore_instructions = [
                ['JNV Attendance Data Restoration Guide', ''],
                ['', ''],
                ['PURPOSE:', 'Use this backup to restore data if database is lost'],
                ['', ''],
                ['SHEET STRUCTURE:', ''],
                ['• "Attendance Data" sheet contains exact database columns', ''],
                ['• First 8 columns match attendance_attendance table exactly', ''],
                ['• Remaining columns are for human readability only', ''],
                ['', ''],
                ['DATABASE COLUMNS (for restoration):', ''],
                ['Column A: id (attendance record ID)', ''],
                ['Column B: student_id (foreign key to student)', ''],
                ['Column C: date (YYYY-MM-DD format)', ''],
                ['Column D: morning_attendance', ''],
                ['Column E: evening_class_attendance', ''],
                ['Column F: morning_pt_attendance', ''],
                ['Column G: games_attendance', ''],
                ['Column H: night_dorm_attendance', ''],
                ['', ''],
                ['REFERENCE COLUMNS (human readable):', ''],
                ['Column I: student_name', ''],
                ['Column J: roll_number', ''],
                ['Column K: class_name', ''],
                ['Column L: house_name', ''],
                ['Column M: branch_name', ''],
                ['', ''],
                ['RESTORATION STEPS:', ''],
                ['1. Export "Attendance Data" sheet to CSV', ''],
                ['2. Keep only columns A-H (database columns)', ''],
                ['3. Import CSV into attendance_attendance table', ''],
                ['4. Ensure student_id exists in attendance_student table', ''],
                ['5. Verify data integrity after import', ''],
                ['', ''],
                ['NOTES:', ''],
                ['• Always backup current data before restoration', ''],
                ['• Verify student_id references are valid', ''],
                ['• Check for duplicate attendance records', ''],
                ['• Test with small batch first', ''],
                [f'• Backup created: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', '']
            ]
            
            # Add restoration instructions
            for row, (instruction, detail) in enumerate(restore_instructions, 1):
                ws_restore.cell(row=row, column=1, value=instruction)
                ws_restore.cell(row=row, column=2, value=detail)
                
                # Apply formatting
                if instruction.endswith(':') or 'JNV Attendance' in instruction:
                    ws_restore.cell(row=row, column=1).font = Font(bold=True, size=12)
                elif instruction.startswith('•') or instruction.startswith('Column'):
                    ws_restore.cell(row=row, column=1).font = Font(size=10)
                elif instruction.isdigit() or instruction.endswith('.'):
                    ws_restore.cell(row=row, column=1).font = Font(bold=True, size=10)
            
            # Auto-adjust restoration sheet column widths
            ws_restore.column_dimensions['A'].width = 45
            ws_restore.column_dimensions['B'].width = 30
            
            # Save workbook
            workbook.save(file_path)
            
            return file_path
            
        except Exception as e:
            print(f"Failed to create Excel file: {str(e)}")
            return None
    
    def _send_backup_email(self, email: str, backup_files: List[str], start_date, end_date):
        """Send Excel backup file via email"""
        subject = f'JNV Attendance Backup - {start_date} to {end_date}'
        
        # Create comprehensive email body
        total_records = Attendance.objects.filter(date__range=[start_date, end_date]).count()
        summary_stats = self.create_summary_report((end_date - start_date).days)
        
        body = f"""
Dear Administrator,

Please find attached the JNV attendance backup in Excel format for the period from {start_date} to {end_date}.

📊 BACKUP SUMMARY:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Date Range: {start_date} to {end_date}
• Total Attendance Records: {total_records:,}
• Total Students Covered: {summary_stats['summary']['total_students']:,}
• Days in Report: {summary_stats['summary']['days_covered']}
• Generated On: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

📁 FILE INFORMATION:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
        for file_path in backup_files:
            filename = os.path.basename(file_path)
            file_size = os.path.getsize(file_path) / 1024  # Size in KB
            body += f"   📊 {filename}\n"
            body += f"   💾 Size: {file_size:.1f} KB\n"
            body += f"   📋 Format: Excel (.xlsx) with Summary Sheet\n"

        body += f"""

🏫 BRANCH BREAKDOWN:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
"""
        for branch_stat in summary_stats['branch_statistics'][:10]:  # Top 10 branches
            branch_name = branch_stat['student__branch__name'] or 'Unknown Branch'
            body += f"   🏢 {branch_name}: {branch_stat['count']:,} records\n"

        body += f"""

📈 EXCEL FILE CONTENTS:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   Sheet 1 - "Attendance Data": Complete attendance records with all fields
   Sheet 2 - "Summary": Statistical overview and branch/class breakdowns
   
   Columns included in main data:
   • Date, Student Name, Roll Number, Class, House, Branch
   • Morning Attendance, Evening Class, Morning PT, Games, Night Dorm

🔒 SECURITY NOTES:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
⚠️  This file contains sensitive student attendance data
🔐 Please store securely and follow institutional data protection policies  
📱 Excel format allows easy viewing, filtering, and analysis
💼 Suitable for administrative records and reporting purposes

This is an automated weekly backup from the JNV Attendance Management System.
If you have any questions or encounter issues, please contact the system administrator.

Best regards,
JNV Attendance Management System
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

System Information:
Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Backup Type: Automated Weekly Backup
Next Backup: {(datetime.now() + timedelta(days=7)).strftime('%Y-%m-%d')}
"""

        try:
            email_message = EmailMessage(
                subject=subject,
                body=body,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=[email],
            )
            
            # Attach Excel file
            for file_path in backup_files:
                email_message.attach_file(file_path)
            
            email_message.send()
            
        except Exception as e:
            print(f"Failed to send email: {str(e)}")
            raise
    
    def _cleanup_files(self, file_paths: List[str]):
        """Clean up temporary files"""
        for file_path in file_paths:
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
            except Exception as e:
                print(f"Failed to cleanup file {file_path}: {str(e)}")


# Convenience functions for easy import
def backup_attendance_data(email: str = 'vinodrc3@gmail.com', days: int = 30) -> bool:
    """Quick function to backup attendance data as Excel"""
    manager = AttendanceBackupManager()
    return manager.create_full_backup(email=email, days_back=days)


def get_attendance_summary(days: int = 7) -> Dict:
    """Quick function to get attendance summary"""
    manager = AttendanceBackupManager()
    return manager.create_summary_report(days_back=days)


def send_weekly_backup(email: str = 'vinodrc3@gmail.com') -> bool:
    """Send weekly Excel backup"""
    manager = AttendanceBackupManager()
    return manager.schedule_weekly_backup(email=email)


def send_daily_backup(email: str = 'vinodrc3@gmail.com') -> bool:
    """Send daily Excel backup"""
    manager = AttendanceBackupManager()
    return manager.schedule_daily_backup(email=email)
